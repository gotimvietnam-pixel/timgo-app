"""
GO TÍM — trích dữ liệu thật từ file Excel báo cáo → sinh file SQL nhập vào Supabase.
Dùng 1 lần (chạy lại được, INSERT dạng upsert).

Nguồn: D:\\canva\\Tim Go\\file bao cao 47\\BAOCAOTHUE2026_07 - Copy.xlsx
Đầu ra: gotim_real_data.sql (cùng thư mục) — Sếp/Antigravity paste vào Supabase SQL Editor.

Quy tắc đã chốt với Sếp (2026-07-07):
- Xóa dữ liệu tài xế mẫu, thay bằng 23 tài xế thật (id = 'tx' + MaID).
- Ghép doanh thu ↔ hồ sơ THEO TÊN (chữ cuối của tên đầy đủ khớp tên viết tắt).
- Bỏ tài xế "TÍNH" (có doanh thu nhưng không có hồ sơ).
- Tài xế không có doanh thu T7 → status 'paused' (tạm nghỉ); có doanh thu → 'active'.
- Chiết khấu mặc định 20% (admin tự chỉnh xuống 10% cho ai được hỗ trợ).

Chạy:
  python build_import_sql.py
"""
import sys
import unicodedata
from pathlib import Path

from openpyxl import load_workbook

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

XLSX = Path(r"D:\canva\Tim Go\file bao cao 47\BAOCAOTHUE2026_07 - Copy.xlsx")
OUT = Path(__file__).resolve().parent / "gotim_real_data.sql"
MONTH = "2026-07"


def sql_str(v):
    """Escape 1 chuỗi cho SQL, trả về 'NULL' nếu rỗng."""
    if v is None:
        return "NULL"
    s = str(v).strip()
    if not s or s == "x":  # 'x' trong Excel = chưa có
        return "NULL"
    return "'" + s.replace("'", "''") + "'"


def last_word_upper(name):
    """Chữ cuối của tên đầy đủ, viết HOA, bỏ dấu để so khớp tên viết tắt."""
    if not name:
        return ""
    return name.strip().split()[-1].upper()


def main():
    if not XLSX.exists():
        print(f"❌ Không thấy file Excel: {XLSX}")
        return
    wb = load_workbook(XLSX, data_only=True)

    # ---- 1. Đọc hồ sơ tài xế (DSTX) ----
    ws = wb["DSTX"]
    rows = list(ws.iter_rows(values_only=True))
    drivers = []
    for r in rows[2:]:
        if not r or r[0] is None:
            continue
        name = (str(r[1]).strip() if r[1] else "")
        if not name:
            continue
        drivers.append({
            "maid": str(r[0]).strip(),
            "name": name,
            "dob": str(r[2]).strip() if r[2] else "",
            "cccd": str(r[3]).strip() if r[3] else "",
            "phone": str(r[4]).strip() if r[4] else "",
            "deposit": str(r[7]).strip() if len(r) > 7 and r[7] else "",
            "uniform": str(r[8]).strip() if len(r) > 8 and r[8] else "",
        })

    # ---- 2. Đọc doanh thu (DANHTHU2026) ----
    ws = wb["DANHTHU2026"]
    rows = list(ws.iter_rows(values_only=True))
    revenue = []
    for r in rows[2:]:
        if not r or r[0] is None:
            continue
        nm = str(r[1]).strip() if r[1] else ""
        if not nm or nm == "0":
            continue
        revenue.append({"abbr": nm.upper(), "rank": r[2], "orders": r[3], "income_k": r[4]})

    # ---- 3. Ghép doanh thu ↔ tài xế THEO TÊN (chữ cuối) ----
    by_lastword = {}
    for d in drivers:
        by_lastword.setdefault(last_word_upper(d["name"]), d)

    matched, unmatched = {}, []
    for rev in revenue:
        d = by_lastword.get(rev["abbr"])
        if d:
            matched[d["maid"]] = rev
        else:
            unmatched.append(rev)  # VD: TÍNH — bỏ qua theo yêu cầu

    # ---- 4. Đọc công nợ chi tiết (CNo) ----
    ws = wb["CNo"]
    rows = list(ws.iter_rows(values_only=True))
    debts = []
    name_index = {last_word_upper(d["name"]): d for d in drivers}
    # cũng index theo tên rút gọn kiểu "S.Kiệt" -> "KIỆT"
    for r in rows[2:]:
        if not r or r[0] is None:
            continue
        ngay_no, kh_no, so_no = r[0], r[1], r[2]
        no_tai = r[4] if len(r) > 4 else None
        ngay_tt = r[5] if len(r) > 5 else None
        if not kh_no or not so_no:
            continue
        # tài xế đầu tiên trong "Võ, S.Kiệt" -> "Võ"
        first_drv = None
        if no_tai:
            first = str(no_tai).split(",")[0].strip()
            key = last_word_upper(first.replace(".", " "))
            first_drv = name_index.get(key)
        debts.append({
            "date": str(ngay_no).strip() if ngay_no else "",
            "cust": str(kh_no).strip(),
            "amount_k": so_no,
            "driver_maid": first_drv["maid"] if first_drv else None,
            "resolved": bool(ngay_tt),
        })

    # ---- 5. Sinh SQL ----
    L = []
    L.append("-- GO TÍM — dữ liệu thật từ Excel báo cáo T7/2026. Chạy trong Supabase SQL Editor.")
    L.append("-- Sinh tự động bởi tools/build_import_sql.py — KIỂM TRA trước khi chạy (có xóa dữ liệu mẫu).\n")

    L.append("-- 5a-0. Bảng vi phạm Zalo (bot audit) — gộp vào đây để chạy 1 lần")
    L.append("create table if not exists zalo_violations (")
    L.append("  id text primary key, zalo_sender_id text not null, driver_name_zalo text,")
    L.append("  driver_id text references users(id) on delete set null,")
    L.append("  violation_type text not null, cuoc_type text, severity text default 'warning',")
    L.append("  detail jsonb, distance_m numeric, speed_kmh numeric,")
    L.append("  occurred_at timestamptz not null, resolved_at timestamptz, resolved_by text,")
    L.append("  created_at timestamptz not null default now());")
    L.append("create index if not exists zalo_violations_driver_id_idx on zalo_violations(driver_id);")
    L.append("create index if not exists zalo_violations_occurred_at_idx on zalo_violations(occurred_at desc);")
    L.append("")

    L.append("-- 5a. Thêm cột hồ sơ tài xế (nếu chưa có)")
    for col in ["dob text", "deposit_note text", "uniform_note text",
                "ao_so text", "ao_size text", "ao_soluong int"]:
        L.append(f"alter table users add column if not exists {col.split()[0]} {col.split()[1]};")
    L.append("create table if not exists driver_monthly_stats (")
    L.append("  id text primary key, driver_id text references users(id) on delete cascade,")
    L.append("  month text not null, orders int default 0, income int default 0, rank int, created_at timestamptz default now());")
    L.append("")

    L.append("-- 5b. Xóa tài xế mẫu (giữ admin), rồi nạp 23 tài xế thật")
    L.append("delete from users where role = 'driver';")
    for d in drivers:
        maid = d["maid"]
        status = "active" if maid in matched else "paused"
        L.append(
            "insert into users (id, name, phone, role, status, cccd, dob, deposit_note, uniform_note, "
            "commission_type, commission_value, wallet, created_at) values ("
            f"'tx{maid}', {sql_str(d['name'])}, {sql_str(d['phone'])}, 'driver', '{status}', "
            f"{sql_str(d['cccd'])}, {sql_str(d['dob'])}, {sql_str(d['deposit'])}, {sql_str(d['uniform'])}, "
            "'percent', 20, 0, now()) on conflict (id) do update set "
            "name=excluded.name, phone=excluded.phone, status=excluded.status, cccd=excluded.cccd, "
            "dob=excluded.dob, deposit_note=excluded.deposit_note, uniform_note=excluded.uniform_note;"
        )
    L.append("")

    L.append(f"-- 5c. Doanh thu tháng {MONTH} (đã ghép theo tên; bỏ tài xế không có hồ sơ)")
    for maid, rev in matched.items():
        inc = int(rev["income_k"] or 0) * 1000
        orders = int(rev["orders"] or 0)
        rank = int(rev["rank"]) if rev["rank"] not in (None, "") else "NULL"
        L.append(
            "insert into driver_monthly_stats (id, driver_id, month, orders, income, rank) values ("
            f"'ms_{MONTH}_tx{maid}', 'tx{maid}', '{MONTH}', {orders}, {inc}, {rank}) "
            "on conflict (id) do update set orders=excluded.orders, income=excluded.income, rank=excluded.rank;"
        )
    L.append("")

    L.append("-- 5d. Công nợ khách (gán tài xế đầu tiên gánh; NULL nếu không khớp tên)")
    for i, db in enumerate(debts):
        amt = int(db["amount_k"] or 0) * 1000
        drv = f"'tx{db['driver_maid']}'" if db["driver_maid"] else "NULL"
        status = "resolved" if db["resolved"] else "pending"
        L.append(
            "insert into debts (id, driver_id, amount, customer_name, status, note, date, created_at) values ("
            f"'exdebt{i}', {drv}, {amt}, {sql_str(db['cust'])}, '{status}', 'Nhập từ Excel T7', "
            f"{sql_str(db['date'])}, now()) on conflict (id) do nothing;"
        )

    OUT.write_text("\n".join(L), encoding="utf-8")

    # ---- 6. Tóm tắt ----
    print(f"✅ Đã sinh: {OUT}")
    print(f"   Tài xế: {len(drivers)} (active {len(matched)}, tạm nghỉ {len(drivers)-len(matched)})")
    print(f"   Doanh thu ghép được: {len(matched)}/{len(revenue)} dòng")
    if unmatched:
        print(f"   ⚠️ Bỏ qua (không có hồ sơ): {', '.join(u['abbr'] for u in unmatched)}")
    no_drv = sum(1 for d in debts if not d["driver_maid"])
    print(f"   Công nợ: {len(debts)} khoản ({no_drv} khoản không khớp tài xế → để NULL)")


if __name__ == "__main__":
    main()
